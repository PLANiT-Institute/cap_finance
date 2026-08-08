#!/usr/bin/env python3
"""Export canonical input tabs from the audit workbook and verify roundtrip parity."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import shutil
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook


SHEET_MAP = {
    "Companies": ("companies.csv", 19),
    "Facilities": ("facilities.csv", 14),
    "Financials": ("company_financials.csv", 9),
    "Technologies": ("technologies.csv", 12),
    "Scenarios": ("scenario_anchors.csv", 10),
    "Scenario_Definitions": ("scenario_definitions.csv", 14),
    "Plans": ("plans.csv", 13),
    "Policy": ("policy_support.csv", 7),
    "Company_Constraints": ("company_constraints.csv", 6),
    "Tech_Constraints": ("technology_constraints.csv", 7),
    "Resource_Constraints": ("resource_constraints.csv", 9),
    "Resource_Benchmarks": ("resource_benchmarks.csv", 16),
    "Transition_Projects": ("transition_projects.csv", 31),
    "Technology_Cost_Evidence": ("technology_cost_evidence.csv", 20),
    "Data_Gaps": ("data_gap_registry.csv", 11),
    "GCAM_Run_Manifest": ("gcam_run_manifest.csv", 25),
    "GCAM_Query_Manifest": ("gcam_query_manifest.csv", 13),
}

AUXILIARY_FILES = (
    "price_process.json",
    "gcam/policy_target_temperature_1p5.xml",
    "gcam/policy_target_temperature_2p0.xml",
)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(65536), b""):
            digest.update(chunk)
    return digest.hexdigest()


def excel_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def normalize(value: str) -> tuple[str, str]:
    text = value.strip()
    if text.lower() in {"true", "false"}:
        return ("bool", text.lower())
    try:
        number = Decimal(text)
    except (InvalidOperation, ValueError):
        return ("text", text)
    return ("number", str(number.normalize()))


def read_csv(path: Path) -> tuple[list[str], list[list[str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))
    return rows[0], rows[1:]


def compare_csv(source: Path, exported: Path) -> tuple[bool, str]:
    source_header, source_rows = read_csv(source)
    export_header, export_rows = read_csv(exported)
    if source_header != export_header:
        return False, "header mismatch"
    if len(source_rows) != len(export_rows):
        return False, f"row-count mismatch: {len(source_rows)} vs {len(export_rows)}"
    for row_index, (left, right) in enumerate(zip(source_rows, export_rows), start=2):
        if len(left) != len(right):
            return False, f"column-count mismatch at row {row_index}"
        for col_index, (a, b) in enumerate(zip(left, right), start=1):
            if normalize(a) != normalize(b):
                return False, f"value mismatch at row {row_index}, column {col_index}: {a!r} vs {b!r}"
    return True, "semantic match"


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--source-dir", type=Path, default=Path("data"))
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--audit-json", type=Path)
    args = parser.parse_args()

    args.output_dir.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(args.workbook, data_only=False, read_only=True)
    audit: dict[str, object] = {
        "workbook": str(args.workbook.resolve()),
        "source_dir": str(args.source_dir.resolve()),
        "output_dir": str(args.output_dir.resolve()),
        "status": "PASS",
        "files": {},
    }

    manifest_rows = []
    for sheet_name, (filename, source_columns) in SHEET_MAP.items():
        sheet = workbook[sheet_name]
        output_path = args.output_dir / filename
        with output_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle, lineterminator="\n")
            for row in sheet.iter_rows(min_row=1, max_col=source_columns, values_only=True):
                writer.writerow([excel_text(value) for value in row])

        ok, detail = compare_csv(args.source_dir / filename, output_path)
        status = "PASS" if ok else "FAIL"
        if not ok:
            audit["status"] = "FAIL"
        file_result = {
            "status": status,
            "detail": detail,
            "source_sha256": sha256(args.source_dir / filename),
            "export_sha256": sha256(output_path),
        }
        audit["files"][filename] = file_result
        manifest_rows.append([filename, sheet_name, status, detail, file_result["source_sha256"], file_result["export_sha256"]])

    for filename in AUXILIARY_FILES:
        source_path = args.source_dir / filename
        output_path = args.output_dir / filename
        output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copyfile(source_path, output_path)
        is_equal = source_path.read_bytes() == output_path.read_bytes()
        status = "PASS" if is_equal else "FAIL"
        if not is_equal:
            audit["status"] = "FAIL"
        audit["files"][filename] = {
            "status": status,
            "detail": "byte match (support file copied unchanged)" if is_equal else "byte mismatch",
            "source_sha256": sha256(source_path),
            "export_sha256": sha256(output_path),
        }
        manifest_rows.append([
            filename,
            "Price_Process" if filename == "price_process.json" else "GCAM_Run_Manifest",
            status,
            audit["files"][filename]["detail"],
            audit["files"][filename]["source_sha256"],
            audit["files"][filename]["export_sha256"],
        ])

    with (args.output_dir / "roundtrip_manifest.csv").open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, lineterminator="\n")
        writer.writerow(["file", "sheet", "status", "detail", "source_sha256", "export_sha256"])
        writer.writerows(manifest_rows)

    audit_path = args.audit_json or args.output_dir.parent / "roundtrip_audit.json"
    audit_path.write_text(json.dumps(audit, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(audit, ensure_ascii=False, indent=2))
    return 0 if audit["status"] == "PASS" else 1


if __name__ == "__main__":
    raise SystemExit(main())
