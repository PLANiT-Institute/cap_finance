#!/usr/bin/env python3
"""엑셀 수집 템플릿 → 분석 입력 CSV 변환.
행 규약: 4행=컬럼명, 5행=설명(제외), 6행~=데이터. 완전 빈 행 제외.
사용: python scripts/xlsx_to_csv.py [--src data/CAP_data_collection_template.xlsx] [--out data/raw]
"""
import argparse, csv, sys
from pathlib import Path
import openpyxl

SHEET_TO_CSV = {
    "source_register": "source_register.csv",
    "D1a_facility_static": "facility_static.csv",
    "D1b_facility_panel": "facility_panel.csv",
    "D2a_scenario_budget": "scenario_budget.csv",
    "D2b_scenario_prices": "scenario_prices.csv",
    "D3_tech_options": "tech_options.csv",
    "D4_price_history": "price_history.csv",
    "D5_policy_support": "policy_support.csv",
    "D6_company_financials": "company_financials.csv",
    "D7_disclosed_plan": "disclosed_plan.csv",
}

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", default="data/CAP_data_collection_template.xlsx")
    ap.add_argument("--out", default="data/raw")
    a = ap.parse_args()
    out = Path(a.out); out.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(a.src, data_only=True)
    for sheet, fname in SHEET_TO_CSV.items():
        if sheet not in wb.sheetnames:
            print(f"[skip] 시트 없음: {sheet}", file=sys.stderr); continue
        ws = wb[sheet]
        header = [c.value for c in ws[4]]
        while header and header[-1] is None: header.pop()
        n = 0
        with open(out/fname, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f); w.writerow(header)
            for row in ws.iter_rows(min_row=6, max_col=len(header), values_only=True):
                if all(v in (None, "") for v in row): continue
                w.writerow(row); n += 1
        print(f"{fname}: {n} rows")

if __name__ == "__main__":
    main()
